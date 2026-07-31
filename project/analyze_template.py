import openpyxl

# Read the Tractor template and write info to a file
with open('templet_analysis.txt', 'w', encoding='utf-8') as f:
    f.write("=== Reading Tractor_Template.xlsx ===\n")
    wb = openpyxl.load_workbook('D:\\PCARDB\\project\\backend\\assets\\templates\\Tractor_Template.xlsx')
    f.write(f"Sheet names: {', '.join(wb.sheetnames)}\n\n")

    if 'PLDMagic' in wb.sheetnames:
        ws = wb['PLDMagic']
        f.write("PLDMagic sheet contents (first 30 rows, columns A-H):\n")
        for row in range(1, 31):
            row_data = []
            for col in range(1, 9):  # A-H
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is None:
                    row_data.append("")
                else:
                    # Convert to string, handling various types
                    try:
                        if isinstance(cell_value, (int, float)):
                            row_data.append(str(cell_value))
                        else:
                            row_data.append(str(cell_value))
                    except:
                        row_data.append("[ENCODE_ERROR]")

            # Format output
            line_parts = []
            for i, cell in enumerate(row_data):
                col_letter = chr(65 + i)  # A, B, C, D, E, F, G, H
                # Truncate long content for readability
                display_val = cell[:20] + "..." if len(cell) > 20 else cell
                line_parts.append(f"{col_letter}:{display_val}")

            f.write(f"Row {row:2d}: {' | '.join(line_parts)}\n")

        # Let's also specifically check the areas mentioned in the code
        f.write("\n" + "="*50 + "\n")
        f.write("SPECIFIC CELL CHECKS FROM LAYOUTS:\n")
        f.write("="*50 + "\n")

        # From the LAYOUTS dictionary in generator.py for TRACTOR:
        tractor_layout = {
            "mobile": (2, 3),      # C2
            "date": (4, 3),        # C4
            "name": (8, 3),        # C8
            "hobli": (13, 3),      # C13
            "village": (14, 3),    # C14
            "taluk": (16, 3),      # C16
            "district": (17, 3),   # C17
            "aadhaar": (31, 3),    # C31
            "dob": (33, 3),        # C33
            "bank_ac": (25, 1),    # A25
            "bank_name": (26, 1)   # A26
        }

        f.write("\nTractor Layout Mapping:\n")
        for field, (row, col) in tractor_layout.items():
            cell_value = ws.cell(row=row, column=col).value
            display_val = "" if cell_value is None else str(cell_value)[:50]
            f.write(f"{field:12} ({row},{col}): '{display_val}'\n")

        # Also check some financial cells mentioned in fill_tractor_details
        f.write("\nFinancial Cells Check:\n")
        financial_cells = [
            (54, 6, "F54 - Tractor Cost"),
            (55, 6, "F55 - Trailer Cost"),
            (56, 6, "F56 - Implement Cost"),
            (54, 8, "H54 - Tractor Bank Loan"),
            (55, 8, "H55 - Trailer Bank Loan"),
            (61, 6, "F61 - Total Project Cost"),
            (62, 6, "F62 - Margin Money"),
            (63, 6, "F63 - Loan Required")
        ]

        for row, col, desc in financial_cells:
            cell_value = ws.cell(row=row, column=col).value
            display_val = "" if cell_value is None else str(cell_value)
            f.write(f"{desc:25} ({row},{col}): {display_val}\n")
    else:
        f.write("PLDMagic sheet not found!\n")

    # Check a few other sheets that might have mapping data
    f.write("\n" + "="*50 + "\n")
    f.write("OTHER SHEETS THAT MIGHT CONTAIN MAPPING:\n")
    f.write("="*50 + "\n")

    interesting_sheets = ['A1', 'A2', 'A3', 'A4', 'B1', 'B2', 'B3', 'B4', 'T1', 'T2', 'T3', 'T4', 'T5', 'T6', 'T7']
    for sheet_name in interesting_sheets:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            f.write(f"\nSheet '{sheet_name}' (first 10 rows, A-D):\n")
            for row in range(1, 11):
                row_data = []
                for col in range(1, 5):  # A-D
                    cell_value = ws.cell(row=row, column=col).value
                    display_val = "" if cell_value is None else str(cell_value)[:20]
                    row_data.append(f"{chr(64+col)}:{display_val}")
                f.write(f"  Row {row:2d}: {' | '.join(row_data)}\n")