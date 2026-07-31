import sys
sys.path.append('backend')

from services.generator import generate_excel
from models import Application, TractorDetails

# Create a minimal test application with ASCII names to avoid encoding issues
app = Application(
    applicant_name_kn="Test Applicant",
    father_name_kn="Test Father",
    age=35,
    gender="Male",
    mobile_no="9876543210",
    aadhaar_no="1234567890123456",
    caste="Hindu",
    farmer_type="Marginal",
    village="Test Village",
    hobli="Test Hobli",
    taluk="Test Taluk",
    district="Test District",
    scheme_type="TRACTOR",
    loan_amount=500000.0
)

# Create tractor details
tractor_details = TractorDetails(
    tractor_make="Mahindra",
    tractor_model="575 DI",
    tractor_hp="45 HP",
    tractor_cost=600000.0,
    tractor_quotation=650000.0,
    tractor_down_payment=150000.0,
    tractor_bank_loan=500000.0,
    trailer_make="Ashok Leyland",
    trailer_cost=150000.0,
    trailer_quotation=180000.0,
    trailer_down_payment=30000.0,
    trailer_bank_loan=150000.0,
    implement_cost=50000.0,
    total_project_cost=800000.0,
    margin_money=300000.0,
    total_loan_amount=500000.0
)

try:
    print("Testing generator with sample data...")
    output_path = generate_excel(app, tractor_details)
    print(f"Generation successful! Output: {output_path}")

    # Check if file exists and get basic info
    import os
    if os.path.exists(output_path):
        file_size = os.path.getsize(output_path)
        print(f"Output file size: {file_size} bytes")

        # Try to read a few key cells to verify data was written correctly
        import openpyxl
        wb = openpyxl.load_workbook(output_path)
        if 'PLDMagic' in wb.sheetnames:
            ws = wb['PLDMagic']

            # Check some key positions
            test_cells = [
                (2, 3, "Mobile (C3)"),
                (4, 3, "Date (C4)"),
                (8, 3, "Name (C8)"),
                (14, 3, "Village (C14)"),
                (16, 3, "Taluk (C16)"),
                (17, 3, "District (C17)"),
                (25, 1, "Bank Account (A25)"),
                (26, 1, "Bank Name (A26)"),
                (54, 6, "Tractor Cost (F54)"),
                (55, 6, "Trailer Cost (F55)"),
                (56, 6, "Implement Cost (F56)"),
                (54, 8, "Tractor Bank Loan (H54)"),
                (55, 8, "Trailer Bank Loan (H55)"),
                (61, 6, "Total Project Cost (F61)"),
                (62, 6, "Margin Money (F62)"),
                (63, 6, "Loan Required (F63)")
            ]

            print("\nVerifying written data in generated file:")
            for row, col, desc in test_cells:
                try:
                    cell_val = ws.cell(row=row, column=col).value
                    display_val = str(cell_val) if cell_val is not None else "(empty)"
                    if len(display_val) > 30:
                        display_val = display_val[:27] + "..."
                    print(f"{desc:25} ({row:2},{col:2}): {display_val}")
                except Exception as e:
                    print(f"{desc:25} ({row:2},{col:2}): Error reading cell - {e}")
        else:
            print("ERROR: PLDMagic sheet not found in generated file")
    else:
        print("ERROR: Output file was not created")

except Exception as e:
    print(f"Error during generation: {e}")
    import traceback
    traceback.print_exc()