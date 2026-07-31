import openpyxl

# Read the Bele chart Excel file
wb = openpyxl.load_workbook('D:\\PCARDB\\project\\Bele chart Temp.xlsx')
sheet = wb.active

# Create a mapping from the Excel data
crop_mapping_from_excel = {}

# Read rows 4-40 to get the crop data
for row in range(4, 41):
    # Get values from columns B, C, D, E
    b_val = sheet.cell(row=row, column=2).value  # Column B (ID/serial)
    c_val = sheet.cell(row=row, column=3).value  # Column C (Crop name in local language)
    d_val = sheet.cell(row=row, column=4).value  # Column D (unit?)
    e_val = sheet.cell(row=row, column=5).value  # Column E (rate/value)

    # Only process rows that have data in column C (crop name)
    if c_val is not None:
        try:
            # Convert values to strings, handling None
            b_str = "" if b_val is None else str(b_val).strip()
            c_str = "" if c_val is None else str(c_val).strip()
            d_str = "" if d_val is None else str(d_val).strip()

            # Try to convert E to integer if it's a number
            try:
                e_int = int(float(e_val)) if e_val is not None else 0
            except:
                e_int = 0

            # Store in our mapping: crop name (local language) -> value
            if c_str and e_int > 0:  # Only store if we have a meaningful crop name and value
                crop_mapping_from_excel[c_str] = e_int

        except Exception as e:
            # Skip rows that cause errors
            pass

# Now let's see what we have
print("CROP MAPPING FROM EXCEL (Bele chart):")
print("=" * 50)
for crop_name, value in sorted(crop_mapping_from_excel.items(), key=lambda x: x[1]):
    print(f"'{crop_name}': {value}")

print("\n\nCURRENT CROP_CHART IN CODE:")
print("=" * 50)
current_crop_chart = {
    'Sugarcane': 11882,
    'Rice': 13793,
    'Jowar': 22120,
    'Maize': 13772,
    'Wheat': 9648,
    'Cotton': 14914,
    'Groundnut': 9722,
    'Sunflower': 22680,
    'Soybean': 7539,
    'Tomato': 78793,
    'Onion': 71685,
    'Chilli': 21000,
    'Banana': 43800,
    'Grapes': 28030,
    'Other': 20000,
}

for crop_name, value in sorted(current_crop_chart.items(), key=lambda x: x[1]):
    print(f"'{crop_name}': {value}")

# Now let's try to map the Kannada names to English names
# Based on visual inspection and common crop names, I'll try to map them
print("\n\nPROPOSED MAPPING BASED ON VISUAL INSPECTION:")
print("=" * 50)

# Let's create a mapping from the Excel Kannada names to English crop names
# Based on the values and common knowledge of crop rates
kannada_to_english = {
    '¨sÀvÀÛ': 'Jowar',                    # Row 5: 11882 (matches Sugarcane in current code?)
    'ªÀÄÄ¸ÀÄQ£ÀeÉÆÃ¼À': 'Rice',           # Row 6: 13793 (matches Rice in current code)
    '¸ÉÃAUÁ': 'Wheat',                   # Row 7: 22120 (matches Jowar in current code?)
    'ºÉå©æqÀ ºÀwÛ': 'Maize',             # Row 8: 13772 (matches Maize in current code)
    '¸ÀÆAiÀÄðPÁAw': 'Sugarcane',         # Row 9: 9648 (matches Wheat in current code?)
    'ºÉå©æqÀ eÉÆÃ¼À': 'Groundnut',       # Row 10: 14914 (close to Cotton? No...)
    'UÉÆÃ¢ü': 'Groundnut',               # Row 11: 9722 (matches Groundnut in current code)
    'vÀA¨ÁPÀÄ': 'Sunflower',             # Row 12: 22680 (matches Sunflower in current code)
    '¢ézÀ¼À zsÁ£Àå': 'Soybean',          # Row 13: 7539 (matches Soybean in current code)
    'PÀ§Äâ': 'Tomato',                   # Row 14: 78793 (matches Tomato in current code)
    'gÉÃµÉä (¸ÁA)': 'Onion',             # Row 16: 71685 (matches Onion in current code)
    'vÀÉAUÀÄ (PÁ¬Ä)': 'Chilli',          # Row 17: 21000 (matches Chilli in current code)
    'ªÀÄiÁªÀÅ': 'Banana',                # Row 18: 43800 (matches Banana in current code)
    'aPÀÄÌ': 'Grapes',                   # Row 19: 28030 (matches Grapes in current code)
    '¨Á¼É': 'Other',                     # Row 20: 79965 (this is high - maybe something else?)
    'zÁ½A¨É': 'Other',                   # Row 21: 31247
    '¥À¥ÁàAiÀÄ': 'Other',                # Row 22: 30378
    '°A¨É': 'Other',                     # Row 23: 22680 (same as Sunflower - duplicate?)
    'zÁæQë (©Ãd gÀ»vÀ)': 'Other',        # Row 25: 165473 (very high - maybe special crop?)
    '¨ÉÆÃgÉ': 'Other',                   # Row 26: 14290
    'UÀÄ¯Á© (¸ÀASÉå)': 'Other',          # Row 27: 51030
    '«¼ÉåzÉ¯É («Ä±Àæ)': 'Other',         # Row 28: 45841
    'PÀ®èAUÀr': 'Other',                 # Row 29: 40703
    'vÀgÀPÁjUÀ¼ÀÄ': 'Other',             # Row 30: 25465
    'Cjó¶t': 'Other',                    # Row 31: 51769
    'Mt ªÉÄ£À¹£ÀPÁ¬Ä': 'Other',          # Row 32: 24146
    'FgÀÄ½î': 'Other',                   # Row 33: 23800
    '¨É¼ÀÄî½î': 'Other',                 # Row 34: 17559
    'D®ÄUÀqÉØ': 'Other',                 # Row 35: 38400
    '¸ÀÉÃªÀAwUÉ': 'Other',               # Row 36: 21107
}

print("Kannada to English mapping attempt:")
for kannada, english in kannada_to_english.items():
    if kannada in crop_mapping_from_excel:
        value = crop_mapping_from_excel[kannada]
        print(f"'{kannada}' -> '{english}': {value}")

print("\n\nCOMPARISON OF VALUES:")
print("=" * 50)
print("Kannada name (from Excel) -> Current English name in code -> Excel value -> Current value")
print("-" * 80)

mismatches = []
for kannada, english in kannada_to_english.items():
    if kannada in crop_mapping_from_excel:
        excel_value = crop_mapping_from_excel[kannada]
        current_value = current_crop_chart.get(english, 0)
        if excel_value != current_value:
            mismatches.append((kannada, english, excel_value, current_value))
            print(f"'{kannada}' -> '{english}': {excel_value} (EXCEL) vs {current_value} (CODE) - MISMATCH")
        else:
            print(f"'{kannada}' -> '{english}': {excel_value} (EXCEL) vs {current_value} (CODE) - MATCH")

print(f"\nTotal mismatches found: {len(mismatches)}")