import openpyxl

wb = openpyxl.load_workbook('backend/assets/templates/Tractor_Template.xlsx')
with open('tractor_template_analysis.txt', 'w', encoding='utf-8') as f:
    f.write("=== Tractor Template Analysis ===\n")
    f.write(f"Sheets: {', '.join(wb.sheetnames)}\n\n")

    if 'PLDMagic' in wb.sheetnames:
        ws = wb['PLDMagic']
        f.write("PLDMagic sheet (first 20 rows, cols A-H):\n")
        for row in range(1, 21):
            row_data = []
            for col in range(1, 9):
                val = ws.cell(row=row, column=col).value
                if val is None:
                    row_data.append("")
                else:
                    # Convert to string, handling various types
                    try:
                        row_data.append(str(val))
                    except:
                        row_data.append("[UNSUPPORTABLE]")

            line = f"Row {row:2d}: "
            for i, cell in enumerate(row_data):
                col_letter = chr(65 + i)  # A, B, C, D, E, F, G, H
                # Truncate long content for readability
                display = cell[:20] + "..." if len(cell) > 20 else cell
                line += f"{col_letter}:{display:<22} | "
            f.write(line.rstrip(" | ") + "\n")

        # Also check some specific areas mentioned in the generator.py LAYOUTS
        f.write("\n\n=== Key Position Analysis (based on generator.py LAYOUTS) ===\n")
        # From generator.py LAYOUTS for TRACTOR:
        tractor_layout = {
            "mobile": (2, 3),      # C2
            "date": (4, 3),        # C4
            "name": (8, 3),        # C8
            "hobli": (13, 3),      # C13
            "village": (14, 3),    # C14
            "taluk": (16, 3),      # C16
            "district": (17, 3),   # C17
            "aadhaar": (31, 3),    # C31
            "dob": (33, 3),        # C33
            "bank_ac": (25, 1),    # A25
            "bank_name": (26, 1)   # A26
        }

        f.write("Current values in template at key positions:\n")
        for field, (row, col) in tractor_layout.items():
            cell_val = ws.cell(row=row, column=col).value
            display_val = str(cell_val) if cell_val is not None else "(empty)"
            f.write(f"{field:12} ({row},{col}): '{display_val}'\n")

        # Check financial cells too
        f.write("\nFinancial cells (from fill_tractor_details function):\n")
        financial_cells = [
            (54, 6, "F54 - Tractor Cost"),
            (55, 6, "F55 - Trailer Cost"),
            (56, 6, "F56 - Implement Cost"),
            (54, 8, "H54 - Tractor Bank Loan"),
            (55, 8, "H55 - Trailer Bank Loan"),
            (61, 6, "F61 - Total Project Cost"),
            (62, 6, "F62 - Margin Money"),
            (63, 6, "F63 - Loan Required")
        ]

        for row, col, desc in financial_cells:
            cell_val = ws.cell(row=row, column=col).value
            display_val = str(cell_val) if cell_val is not None else "(empty)"
            f.write(f"{desc:25} ({row},{col}): {display_val}\n")