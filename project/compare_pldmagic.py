import openpyxl

wb = openpyxl.load_workbook('Temp File.xlsx')
if 'PLDMagic' in wb.sheetnames:
    ws = wb['PLDMagic']
    print("PLDMagic sheet from Temp File.xlsx - Key positions analysis:")
    print("=" * 60)

    # Check the same positions as in the tractor template
    positions_to_check = [
        (2, 3, "Mobile/WhatsApp No (C3)"),
        (4, 3, "Date (C4)"),
        (8, 3, "Name (C8)"),
        (13, 3, "Hobli (C13)"),
        (14, 3, "Village (C14)"),
        (16, 3, "Taluk (C16)"),
        (17, 3, "District (C17)"),
        (31, 3, "Aadhaar (C31)"),
        (33, 3, "DOB (C33)"),
        (25, 1, "Bank Account (A25)"),
        (26, 1, "Bank Name (A26)"),
        (20, 3, "Caste (C20)"),
        (54, 6, "Tractor Cost (F54)"),
        (55, 6, "Trailer Cost (F55)"),
        (56, 6, "Implement Cost (F56)"),
        (54, 8, "Tractor Bank Loan (H54)"),
        (55, 8, "Trailer Bank Loan (H55)"),
        (61, 6, "Total Project Cost (F61)"),
        (62, 6, "Margin Money (F62)"),
        (63, 6, "Loan Required (F63)")
    ]

    for row, col, desc in positions_to_check:
        try:
            cell_val = ws.cell(row=row, column=col).value
            display_val = str(cell_val) if cell_val is not None else "(empty)"
            # Truncate if too long
            if len(display_val) > 30:
                display_val = display_val[:27] + "..."
            print(f"{desc:25} ({row:2},{col:2}): {display_val}")
        except Exception as e:
            print(f"{desc:25} ({row:2},{col:2}): Error - {e}")
else:
    print("No PLDMagic sheet found in Temp File.xlsx")