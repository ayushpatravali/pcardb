import openpyxl

# Let's verify what the generator actually does by looking at the write function
wb = openpyxl.load_workbook('backend/assets/templates/Tractor_Template.xlsx')
ws = wb['PLDMagic']

# Test writing to a cell to see if it works - avoid printing problematic characters
print("Testing write capability...")
cell_c8 = ws.cell(row=8, column=3)
print(f"Before write - C8 type: {type(cell_c8.value)}, is None: {cell_c8.value is None}")
if cell_c8.value is not None:
    print(f"Before write - C8 is string: {isinstance(cell_c8.value, str)}, length: {len(str(cell_c8.value))}")

# Write test data
ws.cell(row=8, column=3, value="TEST NAME")
cell_c8_after = ws.cell(row=8, column=3)
print(f"After write - C8: {cell_c8_after.value}")
print(f"After write - B8: {ws.cell(row=8, column=2).value}")  # Check adjacent cell

# Save test workbook
wb.save('backend/assets/templates/test_write.xlsx')
print("Test write successful - saved as test_write.xlsx")