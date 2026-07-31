"""
Karnataka Co-operative Agricultural & Rural Development Bank
Tractor-Trailer Loan Application PDF Generator
- Reads data from Excel (PLDMagic sheet + calculated sheets)
- Renders all Kannada text correctly using NotoSansKannada font
- Produces a multi-page PDF matching the original form layout
"""

import openpyxl
import datetime
from fpdf import FPDF
from fpdf.enums import XPos, YPos
import os

# ── Font paths ────────────────────────────────────────────────────────────────
FONT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backend", "assets", "fonts")
FONT_REG  = os.path.join(FONT_DIR, "Kannada.ttf")
FONT_BOLD = os.path.join(FONT_DIR, "Kannada.ttf") # Fallback to regular since we don't have bold yet

# ── Helpers ────────────────────────────────────────────────────────────────────
def fmt_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")
    return str(v) if v else ""

def fmt_num(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v)

def load_data(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    def g(sheet, cell):
        v = wb[sheet][cell].value
        if isinstance(v, (datetime.datetime, datetime.date)):
            return fmt_date(v)
        if isinstance(v, float) and v == int(v):
            return int(v)
        return v if v is not None else ""

    d = {}
    # ── PLDMagic master data ──────────────────────────────────────────────────
    d["mobile"]       = fmt_num(g("PLDMagic","C2"))
    d["app_no"]       = fmt_num(g("PLDMagic","C3"))
    d["date"]         = fmt_date(wb["PLDMagic"]["C4"].value)
    d["scheme"]       = g("PLDMagic","C5")
    d["project_cost"] = fmt_num(g("PLDMagic","C6"))
    d["loan_amt"]     = fmt_num(g("PLDMagic","C7"))
    d["applicant"]    = g("PLDMagic","C8")
    d["co_app1"]      = fmt_num(g("PLDMagic","C9"))
    d["co_app2"]      = fmt_num(g("PLDMagic","C10"))
    d["hobli"]        = g("PLDMagic","C13")
    d["village"]      = g("PLDMagic","C14")
    d["post"]         = g("PLDMagic","C15")
    d["taluk"]        = g("PLDMagic","C16")
    d["district"]     = g("PLDMagic","C17")
    d["farmer_type"]  = g("PLDMagic","C18")
    d["loan_type"]    = g("PLDMagic","C19")
    d["caste"]        = g("PLDMagic","C20")
    d["crop"]         = g("PLDMagic","D23")
    d["irrigation"]   = g("PLDMagic","D24")
    d["annual_income"]= fmt_num(g("PLDMagic","D28"))
    d["site_date"]    = fmt_date(wb["PLDMagic"]["D30"].value)
    d["aadhar"]       = g("PLDMagic","D31")
    d["occupation"]   = g("PLDMagic","D32")
    d["dob"]          = fmt_date(wb["PLDMagic"]["D33"].value)
    d["tractor_make"] = g("PLDMagic","C36")
    d["tractor_hp"]   = g("PLDMagic","C37")
    d["tractor_model"]= fmt_num(g("PLDMagic","C38"))
    d["dealer"]       = g("PLDMagic","C39")
    d["trailer_make"] = g("PLDMagic","C42")
    d["trailer_model"]= g("PLDMagic","C43")
    d["trailer_ton"]  = g("PLDMagic","C44")
    d["member_fee"]   = fmt_num(g("PLDMagic","H21"))

    # Land parcels (up to 6 rows, H4:N9 in PLDMagic)
    d["lands"] = []
    for row in range(4, 10):
        sl   = g("PLDMagic", f"H{row}")
        vil  = g("PLDMagic", f"I{row}")
        srv  = g("PLDMagic", f"J{row}")
        area = g("PLDMagic", f"K{row}")
        typ  = g("PLDMagic", f"M{row}")
        rts  = g("PLDMagic", f"N{row}")
        if sl or vil or srv:
            d["lands"].append({
                "sl": fmt_num(sl), "village": vil,
                "survey": fmt_num(srv), "area": fmt_num(area),
                "type": typ, "rights": fmt_num(rts)
            })
    d["total_area"] = fmt_num(g("PLDMagic","K12"))
    d["total_guntas"]= fmt_num(g("PLDMagic","N12"))
    d["land_value"]  = fmt_num(g("PLDMagic","O4"))

    # Project cost breakdown (PLDMagic rows 54-63)
    d["tractor_cost"] = fmt_num(g("PLDMagic","F54"))
    d["trailer_cost"] = fmt_num(g("PLDMagic","F55"))
    d["total_cost"]   = fmt_num(g("PLDMagic","F61"))
    d["own_contrib"]  = fmt_num(g("PLDMagic","F62"))
    d["loan_needed"]  = fmt_num(g("PLDMagic","F63"))

    # Supplementary facilities
    d["tractors_in_village"] = fmt_num(g("PLDMagic","F68"))
    d["hire_demand"]  = g("PLDMagic","F69")
    d["hire_center"]  = g("PLDMagic","F70")
    d["storage_center"]= g("PLDMagic","F72")
    d["fuel_center"]  = g("PLDMagic","F73")
    d["dist_to_land"] = fmt_num(g("PLDMagic","F75"))
    d["dist_to_market"]= fmt_num(g("PLDMagic","F76"))
    d["dist_to_sugar"] = fmt_num(g("PLDMagic","F77"))

    # ── Calculated sheet values ───────────────────────────────────────────────
    d["net_income"]      = fmt_num(g("T3","I24"))
    d["inc_before"]      = fmt_num(g("T5","I2"))
    d["inc_30pct"]       = fmt_num(g("T5","I3"))
    d["inc_after"]       = fmt_num(g("T5","I4"))
    d["valuation_8x"]    = fmt_num(g("T5","I6"))
    d["loan_eligibility"]= fmt_num(g("T5","I8"))
    d["net_loan_elig"]   = fmt_num(g("T5","I12"))
    d["repay_capacity"]  = fmt_num(g("T5","F15"))
    d["annual_instalment"]= fmt_num(g("T5","F18"))
    d["repay_years"]     = fmt_num(g("T5","F20"))
    d["interest_rate"]   = fmt_num(g("T5","F21"))
    d["trailer_income"]  = fmt_num(100960)

    # Tractor usage
    d["hire_income_gross"] = fmt_num(g("T4","E24"))
    d["hire_maint_cost"]   = fmt_num(g("T4","F24"))
    d["hire_income_net"]   = fmt_num(g("T4","G24"))
    d["own_use_cost"]      = fmt_num(g("T4","H24"))
    d["net_hire_income"]   = fmt_num(g("T4","I24"))

    # Economic programme costs
    d["tractor_scheme_cost"] = fmt_num(g("B3","F43"))
    d["trailer_scheme_cost"] = fmt_num(g("B3","F44"))
    d["total_scheme_cost"]   = fmt_num(g("B3","F46"))
    d["own_tractor"]         = fmt_num(g("B3","H43"))
    d["own_trailer"]         = fmt_num(g("B3","H44"))
    d["total_own"]           = fmt_num(g("B3","H46"))
    d["loan_tractor"]        = fmt_num(g("B3","I43"))
    d["loan_trailer"]        = fmt_num(g("B3","I44"))
    d["total_loan_econ"]     = fmt_num(g("B3","I46"))

    return d


# ── PDF Class ─────────────────────────────────────────────────────────────────
class LoanPDF(FPDF):
    def __init__(self, data):
        super().__init__("P", "mm", "A4")
        self.data = data
        self.set_auto_page_break(auto=True, margin=12)
        self.add_font("K",  "",  FONT_REG)
        self.add_font("K",  "B", FONT_BOLD)
        self.add_font("KB", "",  FONT_BOLD)

    # ─── Primitives ───────────────────────────────────────────────────────────
    def kf(self, size=9, bold=False):
        self.set_font("K", "B" if bold else "", size)

    def hdr(self, txt, size=10):
        self.kf(size, bold=True)
        self.cell(0, 7, txt, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")

    def row_label(self, label, value, lw=70, vw=110, h=6):
        self.kf(8.5)
        self.cell(lw, h, label, border=0)
        self.kf(8.5, bold=True)
        self.cell(vw, h, str(value), border=0,
                  new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    def section_title(self, num, txt):
        self.kf(9, bold=True)
        self.cell(10, 6, f"{num})", border=0)
        self.cell(0, 6, txt, border=0, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    def table_header(self, cols):
        self.kf(8, bold=True)
        for label, w in cols:
            self.cell(w, 7, label, border=1, align="C")
        self.ln()

    def table_row(self, vals, widths, bold=False):
        self.kf(8, bold=bold)
        for v, w in zip(vals, widths):
            self.cell(w, 6, str(v), border=1, align="C")
        self.ln()

    def divider(self, h=2):
        self.ln(h)

    def footer_line(self):
        self.kf(8)
        self.cell(0, 5, "\u2500" * 95, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    # ─── Page header ──────────────────────────────────────────────────────────
    def bank_header(self):
        self.kf(10, bold=True)
        self.cell(0, 7,
            "\u00a2 UÉÆÃPÁPÀ vÁ®ÆPÁ ¥ÁæxÀ«ÄPÀ ¸ÀºÀPÁj PÀÈ¶ ªÀÄvÀÄÛ UÁæ«ÄÃt C©üªÀÈ¢Þ",
            new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        self.cell(0, 6,
            "¨ÁåAPÀ ¤AiÀÄ«ÄvÀ, UÉÆÃPÁPÀ. f¯Áè : ¨É¼ÀUÁA«",
            new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        self.divider(2)

    # ─── Page 1: Application form A1 ─────────────────────────────────────────
    def page_application(self):
        d = self.data
        self.add_page()
        self.bank_header()
        self.kf(9, bold=True)
        self.cell(0, 6,
            "\u0cb8\u0cbe\u0cb2\u0ca6 \u0c85\u0cb0\u0ccd\u0c9c\u0cbf  |  \u0cb5\u0cbf\u0cb6\u0cc7\u0cb7 / \u0cb8\u0cbe\u0cae\u0cbe\u0ca8\u0ccd\u0caf / \u0c95\u0cc3\u0cb7\u0cbf / \u0c97\u0ccd\u0cb0\u0cbe\u0cae\u0cc0\u0ca3 / \u0c97\u0cc3\u0cb9 \u0ca8\u0cbf\u0cb0\u0ccd\u0cae\u0cbe\u0ca3 \u0caf\u0ccb\u0c9c\u0ca8\u0cc6",
            new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        self.kf(7.5)
        self.cell(0, 5, "(\u0c85\u0ca8\u0ccd\u0cb5\u0caf\u0cbf\u0cb8\u0ca6\u0cbf\u0cb0\u0cc1\u0cb5\u0cc1\u0ca6\u0ca8\u0ccd\u0ca8\u0cc1 \u0cb9\u0cca\u0ca1\u0cc6\u0ca6\u0cc1 \u0cb9\u0cbe\u0c95\u0cc1\u0cb5\u0cc1\u0ca6\u0cc1)",
            new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        self.divider(2)

        self.kf(8.5)
        self.cell(60, 6, f"\u0cb8\u0cbe\u0cb2\u0ca6 \u0c85\u0cb0\u0ccd\u0c9c\u0cbf \u0cb8\u0c82\u0c96\u0ccd\u0caf\u0cc6 : {d['app_no']}")
        self.cell(60, 6, f"\u0ca6\u0cbf\u0ca8\u0cbe\u0c82\u0c95 : {d['date']}")
        self.cell(0,  6, f"\u0cae\u0cca\u0cac\u0cc8\u0cb2\u0ccd : {d['mobile']}",
            new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        self.cell(60, 6, f"\u0caf\u0ccb\u0c9c\u0ca8\u0cc6 : {d['scheme']}")
        self.cell(60, 6, f"\u0cb8\u0ca6\u0cb8\u0ccd\u0caf\u0ca4\u0ccd\u0cb5 \u0cb6\u0cc1\u0cb2\u0ccd\u0c95 : {d['member_fee']}")
        self.cell(0,  6, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        self.divider(2)

        self.kf(8.5, bold=True)
        self.cell(0, 6, "\u0cae\u0cbe\u0ca8\u0ccd\u0caf\u0cb0\u0cc7,",
            new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        self.kf(8.5)
        self.multi_cell(0, 5.5,
            f"\u0ca8\u0cbe\u0ca8\u0cc1 / \u0ca8\u0cbe\u0cb5\u0cc1 \u0c95\u0cc6\u0cb3\u0c97\u0cc6 \u0ca8\u0cae\u0cc2\u0ca6\u0cbf\u0cb8\u0cbf\u0cb0\u0cc1\u0cb5 \u0c89\u0ca6\u0ccd\u0ca6\u0cc7\u0cb6\u0c95\u0ccd\u0c95\u0cc6  {d['loan_amt']}  "
            "\u0cb0\u0cc2\u0caa\u0cbe\u0caf\u0cbf\u0c97\u0cb3 \u0cb8\u0cbe\u0cb2\u0cb5\u0ca8\u0ccd\u0ca8\u0cc1 \u0ca4\u0cae\u0ccd\u0cae \u0cac\u0ccd\u0caf\u0cbe\u0c82\u0c95\u0cbf\u0ca8\u0cbf\u0c82\u0ca6 \u0c85\u0caa\u0cc7\u0c95\u0ccd\u0cb7\u0cbf\u0cb8\u0cbf "
            "\u0c95\u0cc6\u0cb3\u0c97\u0cbf\u0ca8\u0c82\u0ca4\u0cc6 \u0cb5\u0cbf\u0cb5\u0cb0\u0c97\u0cb3\u0cca\u0c82\u0ca6\u0cbf\u0c97\u0cc6 \u0c88 \u0cb8\u0cbe\u0cb2\u0ca6 \u0c85\u0cb0\u0ccd\u0c9c\u0cbf\u0caf\u0ca8\u0ccd\u0ca8\u0cc1 \u0cb8\u0cb2\u0ccd\u0cb2\u0cbf\u0cb8\u0cbf\u0cb0\u0cc1\u0ca4\u0ccd\u0ca4\u0cc7\u0ca8\u0cc6/\u0cb5\u0cc6.")
        self.divider(2)

        self.section_title("1", "\u0c85\u0cb0\u0ccd\u0c9c\u0cbf\u0ca6\u0cbe\u0cb0 \u0cb5\u0cbf\u0cb5\u0cb0 :")
        self.row_label("\u0cae\u0cc1\u0c96\u0ccd\u0caf \u0c85\u0cb0\u0ccd\u0c9c\u0cbf\u0ca6\u0cbe\u0cb0 \u0cb9\u0cc6\u0cb8\u0cb0\u0cc1", d["applicant"])
        self.row_label("\u0cb8\u0cb9 \u0c85\u0cb0\u0ccd\u0c9c\u0cbf\u0ca6\u0cbe\u0cb0 \u0cb9\u0cc6\u0cb8\u0cb0\u0cc1", d["co_app1"] or "0")
        self.row_label("\u0cb8\u0cb9 \u0c85\u0cb0\u0ccd\u0c9c\u0cbf\u0ca6\u0cbe\u0cb0 \u0cb9\u0cc6\u0cb8\u0cb0\u0cc1", d["co_app2"] or "0")
        self.kf(8.5)
        self.cell(35, 6, "\u0c97\u0ccd\u0cb0\u0cbe\u0cae")
        self.kf(8.5, bold=True)
        self.cell(55, 6, str(d["village"]))
        self.kf(8.5)
        self.cell(20, 6, "\u0c85\u0c82\u0c9a\u0cc6")
        self.kf(8.5, bold=True)
        self.cell(0,  6, str(d["post"]), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        self.kf(8.5)
        self.cell(35, 6, "\u0ca4\u0cbe\u0cb2\u0cc2\u0c95\u0cc1")
        self.kf(8.5, bold=True)
        self.cell(55, 6, str(d["taluk"]))
        self.kf(8.5)
        self.cell(20, 6, "\u0c9c\u0cbf\u0cb2\u0ccd\u0cb2\u0cc6")
        self.kf(8.5, bold=True)
        self.cell(0,  6, str(d["district"]), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        self.divider(2)

        self.section_title("3", "\u0c89\u0ca6\u0ccd\u0ca6\u0cc7\u0cb6\u0cbf\u0ca4 \u0caf\u0ccb\u0c9c\u0ca8\u0cc6 \u0cb5\u0cbf\u0cb5\u0cb0 :")
        cols = [("\u0c95\u0ccd\u0cb0\u0cae \u0cb8\u0c82\u0c96\u0ccd\u0caf\u0cc6",20),("\u0cb8\u0cbe\u0cb2\u0ca6 \u0c89\u0ca6\u0ccd\u0ca6\u0cc7\u0cb6",90),("\u0caf\u0ccb\u0c9c\u0ca8\u0cbe \u0cb5\u0cc6\u0c9a\u0ccd\u0c9a",35),("\u0c85\u0caa\u0cc7\u0c95\u0ccd\u0cb7\u0cbf\u0cb8\u0cbf\u0cb0\u0cc1\u0cb5 \u0cb8\u0cbe\u0cb2\u0ca6 \u0cae\u0cca\u0ca4\u0ccd\u0ca4",35)]
        self.table_header(cols)
        self.table_row(["1", d["scheme"], d["project_cost"], d["loan_amt"]],
                       [20, 90, 35, 35])
        self.table_row(["2", "", "", ""], [20, 90, 35, 35])
        self.table_row(["3", "", "", ""], [20, 90, 35, 35])
        self.divider(3)

    # ─── Build all pages ─────────────────────────────────────────────────────
    def build(self):
        self.page_application()
        # Additional pages can be added here as methods are implemented
        return self


# ── Entry point ───────────────────────────────────────────────────────────────
def generate_from_excel(xlsx_path, output_path=None):
    """Generate PDF from an Excel workbook."""
    data = load_data(xlsx_path)
    pdf = LoanPDF(data)
    pdf.build()
    
    if output_path is None:
        output_path = os.path.splitext(xlsx_path)[0] + "_output.pdf"
    
    pdf.output(output_path)
    print(f"PDF saved to: {output_path}")
    return output_path


def generate_from_dict(app_data, output_path="TEST_OUTPUT.pdf"):
    """Generate PDF from a dictionary (for API use)."""
    # Map API field names to internal names
    d = {
        "mobile": str(app_data.get("mobile_no", "")),
        "app_no": str(app_data.get("application_no", "")),
        "date": str(app_data.get("date", "")),
        "scheme": str(app_data.get("scheme", app_data.get("scheme_type", ""))),
        "project_cost": str(app_data.get("project_cost", app_data.get("total_project_cost", ""))),
        "loan_amt": str(app_data.get("loan_amount", "")),
        "applicant": str(app_data.get("applicant_name_kn", "")),
        "co_app1": str(app_data.get("co_applicant_1", "0")),
        "co_app2": str(app_data.get("co_applicant_2", "0")),
        "hobli": str(app_data.get("hobli", "")),
        "village": str(app_data.get("village", "")),
        "post": str(app_data.get("post", "")),
        "taluk": str(app_data.get("taluk", "")),
        "district": str(app_data.get("district", "")),
        "farmer_type": str(app_data.get("farmer_type", "")),
        "loan_type": str(app_data.get("borrower_type", "")),
        "caste": str(app_data.get("caste", "")),
        "crop": str(app_data.get("current_crop", "")),
        "irrigation": str(app_data.get("irrigation_source", "")),
        "annual_income": str(app_data.get("annual_income", "")),
        "site_date": str(app_data.get("site_date", "")),
        "aadhar": str(app_data.get("aadhaar_no", "")),
        "occupation": str(app_data.get("occupation", "")),
        "dob": str(app_data.get("dob", "")),
        "tractor_make": str(app_data.get("tractor_make", "")),
        "tractor_hp": str(app_data.get("tractor_hp", "")),
        "tractor_model": str(app_data.get("tractor_model", "")),
        "dealer": str(app_data.get("dealer", "")),
        "trailer_make": str(app_data.get("trailer_make", "")),
        "trailer_model": str(app_data.get("trailer_model", "")),
        "trailer_ton": str(app_data.get("trailer_ton", "")),
        "member_fee": str(app_data.get("member_fee", "")),
        "lands": app_data.get("lands", []),
        "total_area": str(app_data.get("total_area", "")),
        "total_guntas": str(app_data.get("total_guntas", "")),
        "land_value": str(app_data.get("land_value", "")),
        "tractor_cost": str(app_data.get("tractor_cost", "")),
        "trailer_cost": str(app_data.get("trailer_cost", "")),
        "total_cost": str(app_data.get("total_cost", "")),
        "own_contrib": str(app_data.get("own_contrib", app_data.get("margin_money", ""))),
        "loan_needed": str(app_data.get("loan_needed", app_data.get("loan_amount", ""))),
        "age": str(app_data.get("age", "")),
        "mobile_no": str(app_data.get("mobile_no", "")),
        # Defaults for missing calculated fields
        "tractors_in_village": "", "hire_demand": "", "hire_center": "",
        "storage_center": "", "fuel_center": "", "dist_to_land": "",
        "dist_to_market": "", "dist_to_sugar": "",
        "net_income": "", "inc_before": "", "inc_30pct": "", "inc_after": "",
        "valuation_8x": "", "loan_eligibility": "", "net_loan_elig": "",
        "repay_capacity": "", "annual_instalment": "", "repay_years": "",
        "interest_rate": "", "trailer_income": "",
        "hire_income_gross": "", "hire_maint_cost": "", "hire_income_net": "",
        "own_use_cost": "", "net_hire_income": "",
        "tractor_scheme_cost": "", "trailer_scheme_cost": "", "total_scheme_cost": "",
        "own_tractor": "", "own_trailer": "", "total_own": "",
        "loan_tractor": "", "loan_trailer": "", "total_loan_econ": "",
    }
    
    pdf = LoanPDF(d)
    pdf.build()
    pdf.output(output_path)
    print(f"PDF saved to: {output_path}")
    return output_path


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        generate_from_excel(sys.argv[1])
    else:
        # Test with dummy data reflecting the new schema
        test_data = {
            "application_no": "PCARDB-2026-001",
            "applicant_name_kn": "\u0cb0\u0cae\u0cc7\u0cb6 \u0c95\u0cc1\u0cae\u0cbe\u0cb0 \u0caa\u0cbe\u0c9f\u0cc0\u0cb2",
            "co_applicant_1": "\u0cb8\u0cc1\u0ca8\u0cbf\u0ca4\u0cbe \u0caa\u0cbe\u0c9f\u0cc0\u0cb2 (Spouse)",
            "co_applicant_2": "\u0cb0\u0cbe\u0c9c\u0cc1 \u0caa\u0cbe\u0c9f\u0cc0\u0cb2 (Son)",
            "village": "\u0cac\u0cc6\u0cb3\u0c97\u0cbe\u0cb5\u0cbf",
            "taluk": "\u0c97\u0ccb\u0c95\u0cbe\u0c95",
            "district": "\u0cac\u0cc6\u0cb3\u0c97\u0cbe\u0cb5\u0cbf",
            "mobile_no": "9876543210",
            "scheme_type": "\u0c9f\u0ccd\u0cb0\u0cbe\u0c95\u0ccd\u0c9f\u0cb0\u0ccd \u0caf\u0ccb\u0c9c\u0ca8\u0cc6",
            "caste": "\u0cb2\u0cbf\u0c82\u0c97\u0cbe\u0caf\u0ca4",
            "age": "35",
            "date": "23/05/2026",
            "farmer_type": "\u0ca6\u0cca\u0ca1\u0ccd\u0ca1 \u0cb0\u0cc8\u0ca4",
            "current_crop": "\u0c95\u0cac\u0ccd\u0cac\u0cc1 (Sugarcane)",
            "irrigation_source": "\u0c95\u0cca\u0cb3\u0cb5\u0cc6 \u0cac\u0cbe\u0cb5\u0cbf (Borewell)",
            "total_cost": "750000",
            "loan_amount": "500000",
            "margin_money": "250000",
            "total_area": "4",
            "total_guntas": "20",
            "lands": [
                {"sl": "1", "village": "\u0cac\u0cc6\u0cb3\u0c97\u0cbe\u0cb5\u0cbf", "survey": "123/A", "area": "2", "rights": "Yes"},
                {"sl": "2", "village": "\u0cac\u0cc6\u0cb3\u0c97\u0cbe\u0cb5\u0cbf", "survey": "123/B", "area": "2.2", "rights": "Yes"}
            ]
        }
        generate_from_dict(test_data, "TEST_OUTPUT.pdf")
