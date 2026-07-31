import openpyxl

# Read the Bele chart Excel file
wb = openpyxl.load_workbook('D:\\PCARDB\\project\\Bele chart Temp.xlsx')
sheet = wb.active

# Write to a file to avoid console encoding issues
with open('bele_chart_data.txt', 'w', encoding='utf-8') as f:
    f.write("Bele Chart Data:\n")
    f.write("="*50 + "\n")

    # Find the actual data area - let's check rows 4-40
    for row in range(4, 41):
        # Get values from columns B, C, D, E (since column A seems empty)
        b_val = sheet.cell(row=row, column=2).value  # Column B
        c_val = sheet.cell(row=row, column=3).value  # Column C
        d_val = sheet.cell(row=row, column=4).value  # Column D
        e_val = sheet.cell(row=row, column=5).value  # Column E

        # Only process rows that have data in column C (crop name)
        if c_val is not None:
            try:
                b_str = "" if b_val is None else str(b_val).strip()
                c_str = "" if c_val is None else str(c_val).strip()
                d_str = "" if d_val is None else str(d_val).strip()
                e_str = "" if e_val is None else str(e_val).strip()

                # Try to convert E to integer if it's a number
                try:
                    e_int = int(float(e_val)) if e_val is not None else 0
                except:
                    e_int = 0

                f.write(f"Row {row:2d}: B='{b_str}' | C='{c_str}' | D='{d_str}' | E={e_int}\n")
            except Exception as e:
                f.write(f"Row {row:2d}: ERROR processing row - {e}\n")