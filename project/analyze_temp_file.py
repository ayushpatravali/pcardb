import openpyxl

print("=== Analyzing Temp File.xlsx ===")
wb = openpyxl.load_workbook('Temp File.xlsx')
sheet = wb.active

print(f"Sheet name: {sheet.title}")
print(f"Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")

# Let's look at a sample of the data to understand the structure
print("\nFirst 15 rows, columns A-H:")
for row in range(1, 16):
    row_data = []
    for col in range(1, 9):  # A-H
        cell_value = sheet.cell(row=row, column=col).value
        if cell_value is None:
            row_data.append("")
        else:
            # Truncate long values for display
            str_val = str(cell_value)
            if len(str_val) > 20:
                str_val = str_val[:17] + "..."
            row_data.append(str_val)
    print(f"Row {row:2}: {' | '.join(f'{cell:<12}' for cell in row_data)}")

# Check if there's a PLDMagic sheet
if 'PLDMagic' in wb.sheetnames:
    print("\\nFound PLDMagic sheet in Temp File.xlsx")
    ws_pld = wb['PLDMagic']
    print(f"PLDMagic dimensions: {ws_pld.max_row} rows x {ws_pld.max_column} columns")
else:
    print("\nNo PLDMagic sheet in Temp File.xlsx")
    print("Available sheets:", wb.sheetnames)