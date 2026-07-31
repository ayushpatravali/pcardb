import openpyxl

# Read the Tractor template
print("=== Reading Tractor_Template.xlsx ===")
wb = openpyxl.load_workbook('D:\\PCARDB\\project\\backend\\assets\\templates\\Tractor_Template.xlsx')
print(f"Sheet names: {wb.sheetnames}")

if 'PLDMagic' in wb.sheetnames:
    ws = wb['PLDMagic']
    print("\nPLDMagic sheet contents (first 20 rows, columns A-H):")
    for row in range(1, 21):
        row_data = []
        for col in range(1, 9):  # A-H
            cell_value = ws.cell(row=row, column=col).value
            # Handle None values and convert to string safely
            if cell_value is None:
                row_data.append("")
            else:
                # Try to convert to string, handling potential encoding issues
                try:
                    row_data.append(str(cell_value))
                except:
                    row_data.append("[ENCODING ERROR]")
        # Print with limited width to avoid terminal issues
        formatted_cells = []
        for cell in row_data:
            if len(cell) > 15:
                formatted_cells.append(cell[:12] + "...")
            else:
                formatted_cells.append(cell.ljust(15))
        print(f"Row {row:2d}: {' | '.join(formatted_cells)}")

# Also check if there are other sheets that might contain mapping data
print("\n" + "="*50)
print("Checking other sheets for mapping data...")
for sheet_name in wb.sheetnames:
    if sheet_name != 'PLDMagic':
        ws = wb[sheet_name]
        print(f"\nSheet '{sheet_name}' (first 5 rows, first 5 columns):")
        for row in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(6, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is None:
                    row_data.append("")
                else:
                    try:
                        row_data.append(str(cell_value))
                    except:
                        row_data.append("[ENCODING ERROR]")
            formatted_cells = []
            for cell in row_data:
                if len(cell) > 15:
                    formatted_cells.append(cell[:12] + "...")
                else:
                    formatted_cells.append(cell.ljust(15))
            print(f"  Row {row}: {' | '.join(formatted_cells)}")
else:
    print("PLDMagic sheet not found!")